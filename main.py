from openpyxl import load_workbook

# Carregando o arquivo Excel
wb = load_workbook('teste.xlsx')

# Selecionar apenas a primeira folha
sheet = wb.active

# Printar o valor da primeira coluna da primeira linha (se for preencher com mais usa o loop lá de baixo)
print("O valor presente na primeira linha da primeira coluna é: ", sheet['A1'].value)

# Printar o número de linhas e colunas (Opcional)
print('Temos esse número de linhas: ', sheet.max_row)
print('Temos esse número de colunas', sheet.max_column)

# Printar todos os nomes da coluna de índice 0
'''
for col in sheet.iter_cols(values_only=True):
    print(col[0])
'''
